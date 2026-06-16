from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO


def generate_excel(bookings):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    # Header row
    headers = ["ID", "Name", "Contact", "Service Type", "Preferred Time",
               "Notes", "Status", "Booked On"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for b in bookings:
        ws.append([
            b.id,
            b.name,
            b.contact,
            b.service_type,
            b.preferred_time,
            b.notes,
            b.status,
            b.created_at.strftime('%d-%m-%Y %H:%M')
        ])

    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output